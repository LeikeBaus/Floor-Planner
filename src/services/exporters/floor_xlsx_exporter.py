"""XLSX exporter for floor data with formatted spreadsheets."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.floor import Floor


class FloorXlsxExporter:
    """Export floor and room data to Excel workbook with formatting."""

    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def export(self, floor: Floor, output_path: Path) -> None:
        """Generate Excel workbook with floor summary, rooms, and walls data."""
        target = Path(output_path)
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        self._add_summary_sheet(workbook, floor)
        self._add_rooms_sheet(workbook, floor)
        self._add_walls_sheet(workbook, floor)
        self._add_windows_sheet(workbook, floor)
        self._add_doors_sheet(workbook, floor)
        self._add_openings_sheet(workbook, floor)
        self._add_stairs_sheet(workbook, floor)

        workbook.save(target)

    def _add_summary_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add floor summary sheet."""
        sheet = workbook.create_sheet("Summary", 0)

        summary_data: list[list[object]] = [
            ["Floor Name", floor.name],
            ["Elevation (mm)", floor.elevation],
            [
                "Floor Area Total (m²)",
                f"{floor.floor_area_total / 1_000_000.0:.2f}",
            ],
            [
                "Living Area Total (m²)",
                f"{floor.living_area_total / 1_000_000.0:.2f}",
            ],
            ["Room Count", len(floor.rooms)],
            ["Wall Count", len(floor.walls)],
            ["Window Count", len(floor.windows)],
            ["Door Count", len(floor.doors)],
            ["Opening Count", len(floor.openings)],
            ["Stair Count", len(floor.stairs)],
        ]

        for row_idx, (key, value) in enumerate(summary_data, 1):
            sheet[f"A{row_idx}"] = key
            sheet[f"B{row_idx}"] = value
            sheet[f"A{row_idx}"].font = Font(bold=True)

        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 20

    def _add_rooms_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add rooms sheet."""
        if not floor.rooms:
            return

        sheet = workbook.create_sheet("Rooms")

        headers = [
            "Room Name",
            "Floor Area (m²)",
            "Living Area (m²)",
            "Color",
            "Point Count",
        ]
        self._write_headers(sheet, headers)

        for row_idx, room in enumerate(floor.rooms, 2):
            room_floor_area_m2 = room.floor_area / 1_000_000.0
            room_living_area_m2 = room.living_area / 1_000_000.0
            point_count = len(room.polygon)

            sheet[f"A{row_idx}"] = room.name
            sheet[f"B{row_idx}"] = room_floor_area_m2
            sheet[f"C{row_idx}"] = room_living_area_m2
            sheet[f"D{row_idx}"] = room.color
            sheet[f"E{row_idx}"] = point_count

            sheet[f"B{row_idx}"].number_format = "0.00"
            sheet[f"C{row_idx}"].number_format = "0.00"

        self._set_column_widths(sheet, [25, 20, 20, 15, 15])

    def _add_walls_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add walls sheet."""
        if not floor.walls:
            return

        sheet = workbook.create_sheet("Walls")

        headers = [
            "Wall ID",
            "Start X (mm)",
            "Start Y (mm)",
            "End X (mm)",
            "End Y (mm)",
            "Length (m)",
            "Type",
        ]
        self._write_headers(sheet, headers)

        for row_idx, wall in enumerate(floor.walls, 2):
            length_m = wall.length / 1000.0

            sheet[f"A{row_idx}"] = wall.id[:8]
            sheet[f"B{row_idx}"] = wall.start.x
            sheet[f"C{row_idx}"] = wall.start.y
            sheet[f"D{row_idx}"] = wall.end.x
            sheet[f"E{row_idx}"] = wall.end.y
            sheet[f"F{row_idx}"] = length_m
            sheet[f"G{row_idx}"] = str(wall.wall_type)

            sheet[f"F{row_idx}"].number_format = "0.00"

        self._set_column_widths(sheet, [12, 15, 15, 15, 15, 12, 15])

    def _add_windows_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add windows sheet."""
        if not floor.windows:
            return

        sheet = workbook.create_sheet("Windows")

        headers = [
            "Window ID",
            "Wall ID",
            "Position (mm)",
            "Width (mm)",
            "Height (mm)",
        ]
        self._write_headers(sheet, headers)

        for row_idx, window in enumerate(floor.windows, 2):
            sheet[f"A{row_idx}"] = window.id[:8]
            sheet[f"B{row_idx}"] = window.wall_id[:8]
            sheet[f"C{row_idx}"] = window.position
            sheet[f"D{row_idx}"] = window.width
            sheet[f"E{row_idx}"] = window.height

        self._set_column_widths(sheet, [12, 12, 15, 12, 12])

    def _add_doors_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add doors sheet."""
        if not floor.doors:
            return

        sheet = workbook.create_sheet("Doors")

        headers = [
            "Door ID",
            "Wall ID",
            "Position (mm)",
            "Width (mm)",
            "Swing Direction",
        ]
        self._write_headers(sheet, headers)

        for row_idx, door in enumerate(floor.doors, 2):
            sheet[f"A{row_idx}"] = door.id[:8]
            sheet[f"B{row_idx}"] = door.wall_id[:8]
            sheet[f"C{row_idx}"] = door.position
            sheet[f"D{row_idx}"] = door.width
            sheet[f"E{row_idx}"] = door.swing_direction

        self._set_column_widths(sheet, [12, 12, 15, 12, 18])

    def _add_openings_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add openings sheet."""
        if not floor.openings:
            return

        sheet = workbook.create_sheet("Openings")

        headers = [
            "Opening ID",
            "Wall ID",
            "Position (mm)",
            "Width (mm)",
        ]
        self._write_headers(sheet, headers)

        for row_idx, opening in enumerate(floor.openings, 2):
            sheet[f"A{row_idx}"] = opening.id[:8]
            sheet[f"B{row_idx}"] = opening.wall_id[:8]
            sheet[f"C{row_idx}"] = opening.position
            sheet[f"D{row_idx}"] = opening.width

        self._set_column_widths(sheet, [12, 12, 15, 12])

    def _add_stairs_sheet(self, workbook: Workbook, floor: Floor) -> None:
        """Add stairs sheet."""
        if not floor.stairs:
            return

        sheet = workbook.create_sheet("Stairs")

        headers = [
            "Stair ID",
            "Position X (mm)",
            "Position Y (mm)",
            "Width (mm)",
            "Depth (mm)",
            "Type",
            "Orientation (deg)",
        ]
        self._write_headers(sheet, headers)

        for row_idx, stair in enumerate(floor.stairs, 2):
            sheet[f"A{row_idx}"] = stair.id[:8]
            sheet[f"B{row_idx}"] = stair.position_x
            sheet[f"C{row_idx}"] = stair.position_y
            sheet[f"D{row_idx}"] = stair.width
            sheet[f"E{row_idx}"] = stair.depth
            sheet[f"F{row_idx}"] = stair.stair_type
            sheet[f"G{row_idx}"] = stair.orientation_degrees

        self._set_column_widths(sheet, [12, 15, 15, 12, 12, 15, 12])

    def _write_headers(
        self, sheet: Worksheet, headers: list[str]
    ) -> None:
        """Write headers with formatting."""
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGNMENT

    def _set_column_widths(self, sheet: Worksheet, widths: list[int]) -> None:
        """Set column widths."""
        for col_idx, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
